"""Build a colour-coded Excel workbook for a search run's jobs."""
import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    'Job Title', 'Company', 'Location', 'Date Posted', 'Employment Type',
    'Seniority Level', 'Salary', 'Sponsorship Flag', 'Match Score',
    'Match Reason', 'Application Link', 'Tailored CV Filename',
    'Job Required Skills', 'Missing Skills', 'ATS Score',
]

# 1-based index of the Match Score column (for colour coding).
_SCORE_COL = HEADERS.index('Match Score') + 1

_HEADER_FILL = PatternFill('solid', fgColor='D9D9D9')
_GREEN = PatternFill('solid', fgColor='C6EFCE')
_YELLOW = PatternFill('solid', fgColor='FFEB9C')
_RED = PatternFill('solid', fgColor='FFC7CE')


def _score_fill(score):
    if score is None:
        return None
    if score >= 75:
        return _GREEN
    if score >= 50:
        return _YELLOW
    return _RED


def _job_row(job):
    tailored_name = ''
    if job.tailored_pdf:
        tailored_name = os.path.basename(job.tailored_pdf.name)
    return [
        job.title,
        job.company,
        job.location,
        job.date_posted,
        job.employment_type,
        job.seniority_level,
        job.salary,
        job.get_sponsorship_flag_display(),
        job.match_score if job.match_score is not None else '',
        job.match_reason,
        job.application_link,
        tailored_name,
        ', '.join(job.job_skills or []),
        ', '.join(job.missing_skills or []),
        job.ats_score if job.ats_score is not None else '',
    ]


def build_workbook(search_run):
    """Return a BytesIO containing the formatted .xlsx for ``search_run``."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Jobs'

    # Header row.
    header_font = Font(bold=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical='center')

    # Data rows.
    for row_idx, job in enumerate(search_run.jobs.all().order_by('-match_score'), start=2):
        for col, value in enumerate(_job_row(job), start=1):
            ws.cell(row=row_idx, column=col, value=value)
        fill = _score_fill(job.match_score)
        if fill is not None:
            ws.cell(row=row_idx, column=_SCORE_COL).fill = fill

    # Freeze header, enable autofilter, set sensible column widths.
    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'
    widths = [30, 24, 22, 14, 16, 16, 18, 16, 12, 45, 40, 32, 30, 30, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
